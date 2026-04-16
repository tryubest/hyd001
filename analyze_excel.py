import openpyxl
from collections import defaultdict

EXCEL_PATH = r'c:\Users\huangyuandong\Desktop\EXPORT_20260416_161509.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

rows_data = []
for row_idx in range(2, ws.max_row + 1):
    a_val = ws.cell(row_idx, 1).value
    b_val = ws.cell(row_idx, 2).value
    order_id = str(a_val).strip() if a_val and str(a_val).strip() else str(b_val).strip()
    row = {
        'row_idx': row_idx,
        'order': order_id,
        'material': ws.cell(row_idx, 3).value,
        'priority': int(ws.cell(row_idx, 4).value),
        'demand': float(ws.cell(row_idx, 6).value or 0),
        'group': str(ws.cell(row_idx, 7).value),
        'excel_L': float(ws.cell(row_idx, 12).value or 0),
        'excel_M': float(ws.cell(row_idx, 13).value or 0),
        'excel_O': float(ws.cell(row_idx, 15).value or 0),
        'excel_P': float(ws.cell(row_idx, 16).value or 0),
        'excel_Q': float(ws.cell(row_idx, 17).value or 0),
        'init_inv': float(ws.cell(row_idx, 9).value or 0),
        'init_po': float(ws.cell(row_idx, 10).value or 0),
    }
    rows_data.append(row)

print("=== Excel 数据统计概览 ===")
unique_orders = set(r["order"] for r in rows_data)
unique_mats = set(r["material"] for r in rows_data)
unique_grps = set(r["group"] for r in rows_data)
print(f"总行数: {len(rows_data)}")
print(f"唯一工单数: {len(unique_orders)}")
print(f"唯一物料数: {len(unique_mats)}")
print(f"唯一替代料组数: {len(unique_grps)}")
print()

total_demand = sum(r["demand"] for r in rows_data)
total_inv_used = sum(r["excel_L"] for r in rows_data)
total_po_used = sum(r["excel_M"] for r in rows_data)
total_shortage = sum(r["excel_Q"] for r in rows_data)

print("=== 供需汇总 ===")
print(f"总需求数量: {total_demand:,.0f}")
print(f"库存分配总量(占用库存): {total_inv_used:,.0f}")
print(f"PO分配总量(占用PO): {total_po_used:,.0f}")
print(f"总欠数(实际净需求): {total_shortage:,.0f}")
if total_demand > 0:
    print(f"满足率: {(total_inv_used + total_po_used) / total_demand * 100:.2f}%")
print()

shortage_rows = [r for r in rows_data if r["excel_Q"] > 0]
print(f"=== 有欠数的工单明细 ({len(shortage_rows)} 条) ===")
shortage_by_order = defaultdict(list)
for r in shortage_rows:
    shortage_by_order[r["order"]].append(r)

for order in sorted(shortage_by_order.keys()):
    items = shortage_by_order[order]
    total_q = sum(x["excel_Q"] for x in items)
    print(f"  工单 {order}: 总欠数 = {total_q:,.0f}")
    for item in items:
        print(f"    物料={item['material']}, 替代组={item['group']}, "
              f"需求={item['demand']:,.0f}, 欠数={item['excel_Q']:,.0f}")

print()
print("=== 分配来源统计 ===")
pri1_demands = [r for r in rows_data if r["priority"] == 1 and r["demand"] > 0]
inv_only = sum(1 for r in pri1_demands if r["excel_Q"] == 0 and r["excel_L"] > 0 and r["excel_M"] == 0)
po_only = sum(1 for r in pri1_demands if r["excel_Q"] == 0 and r["excel_L"] == 0 and r["excel_M"] > 0)
inv_and_po = sum(1 for r in pri1_demands if r["excel_Q"] == 0 and r["excel_L"] > 0 and r["excel_M"] > 0)
shortage_cnt = sum(1 for r in pri1_demands if r["excel_Q"] > 0)
no_own_alloc = sum(1 for r in pri1_demands if r["excel_L"] == 0 and r["excel_M"] == 0)
sub_served = sum(1 for r in rows_data if r["priority"] > 1 and (r["excel_L"] > 0 or r["excel_M"] > 0))

print(f"  需求行(优先级1)总数: {len(pri1_demands)}")
print(f"    仅由库存满足: {inv_only}")
print(f"    由库存+PO满足: {inv_and_po}")
print(f"    仅由PO满足: {po_only}")
print(f"    有欠数(未完全满足): {shortage_cnt}")
print(f"    本物料无分配(由替代料或完全短缺): {no_own_alloc}")
print(f"  替代料参与分配的行数: {sub_served}")
