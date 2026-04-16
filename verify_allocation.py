import openpyxl
from collections import defaultdict

EXCEL_PATH = r'c:\Users\huangyuandong\Desktop\EXPORT_20260416_161509.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

rows_data = []
for row_idx in range(2, ws.max_row + 1):
    a_val = ws.cell(row_idx, 1).value  # 生产订单
    b_val = ws.cell(row_idx, 2).value  # 计划订单
    order_id = str(a_val).strip() if a_val and str(a_val).strip() else str(b_val).strip()

    row = {
        'row_idx': row_idx,
        'order': order_id,
        'material': ws.cell(row_idx, 3).value,
        'priority': int(ws.cell(row_idx, 4).value),
        'date': ws.cell(row_idx, 5).value,
        'demand': float(ws.cell(row_idx, 6).value or 0),
        'group': str(ws.cell(row_idx, 7).value),
        'prob': float(ws.cell(row_idx, 8).value or 0),
        'init_inv': float(ws.cell(row_idx, 9).value or 0),
        'init_po': float(ws.cell(row_idx, 10).value or 0),
        'excel_L': float(ws.cell(row_idx, 12).value or 0),
        'excel_M': float(ws.cell(row_idx, 13).value or 0),
        'excel_N': ws.cell(row_idx, 14).value or '',
        'excel_O': float(ws.cell(row_idx, 15).value or 0),
        'excel_P': float(ws.cell(row_idx, 16).value or 0),
        'excel_Q': float(ws.cell(row_idx, 17).value or 0),
    }
    rows_data.append(row)

mat_initial_inv = {}
mat_initial_po = {}
for r in rows_data:
    mat = r['material']
    if r['init_inv'] > 0:
        if mat not in mat_initial_inv:
            mat_initial_inv[mat] = r['init_inv']
    if r['init_po'] > 0:
        if mat not in mat_initial_po:
            mat_initial_po[mat] = r['init_po']

mat_remaining_inv = {m: v for m, v in mat_initial_inv.items()}
mat_remaining_po = {m: v for m, v in mat_initial_po.items()}

order_groups = []
i = 0
while i < len(rows_data):
    current_order = rows_data[i]['order']
    current_group = rows_data[i]['group']
    group_rows = []
    while i < len(rows_data) and rows_data[i]['order'] == current_order and rows_data[i]['group'] == current_group:
        group_rows.append(rows_data[i])
        i += 1
    order_groups.append(group_rows)

errors = []
total_groups = len(order_groups)
correct_groups = 0

for grp_rows in order_groups:
    grp_rows_sorted = sorted(grp_rows, key=lambda x: x['priority'])

    total_demand = sum(r['demand'] for r in grp_rows_sorted)

    remaining_demand = total_demand

    computed_L = {}
    computed_M = {}

    for r in grp_rows_sorted:
        mat = r['material']
        avail = mat_remaining_inv.get(mat, 0)
        alloc = min(avail, remaining_demand)
        computed_L[r['row_idx']] = alloc
        remaining_demand -= alloc

    for r in grp_rows_sorted:
        mat = r['material']
        avail = mat_remaining_po.get(mat, 0)
        alloc = min(avail, remaining_demand)
        computed_M[r['row_idx']] = alloc
        remaining_demand -= alloc

    shortage = remaining_demand

    for r in grp_rows_sorted:
        mat = r['material']
        l_val = computed_L[r['row_idx']]
        m_val = computed_M[r['row_idx']]
        mat_remaining_inv[mat] = mat_remaining_inv.get(mat, 0) - l_val
        mat_remaining_po[mat] = mat_remaining_po.get(mat, 0) - m_val

    computed_O = {}
    computed_P = {}
    computed_Q = {}
    for r in grp_rows_sorted:
        mat = r['material']
        computed_O[r['row_idx']] = mat_remaining_inv.get(mat, 0)
        computed_P[r['row_idx']] = mat_remaining_po.get(mat, 0)
        if r['priority'] == min(x['priority'] for x in grp_rows_sorted):
            computed_Q[r['row_idx']] = shortage
        else:
            computed_Q[r['row_idx']] = 0

    has_error = False
    for r in grp_rows_sorted:
        rid = r['row_idx']
        checks = [
            ('L(占用库存)', r['excel_L'], computed_L[rid]),
            ('M(占用PO)', r['excel_M'], computed_M[rid]),
            ('O(剩余库存)', r['excel_O'], computed_O[rid]),
            ('P(剩余PO)', r['excel_P'], computed_P[rid]),
            ('Q(实际净需求)', r['excel_Q'], computed_Q[rid]),
        ]
        for col_name, excel_val, computed_val in checks:
            if abs(excel_val - computed_val) > 0.01:
                has_error = True
                errors.append({
                    'row': rid,
                    'order': r['order'],
                    'material': r['material'],
                    'priority': r['priority'],
                    'group': r['group'],
                    'column': col_name,
                    'excel_value': excel_val,
                    'computed_value': computed_val,
                    'diff': excel_val - computed_val,
                    'demand': r['demand'],
                })

    if not has_error:
        correct_groups += 1

print(f"=== 分配验证结果 ===")
print(f"总共验证: {total_groups} 个工单-替代料组")
print(f"正确: {correct_groups} 个")
print(f"有差异: {total_groups - correct_groups} 个")
print(f"差异明细条数: {len(errors)} 条")
print()

if errors:
    print("=== 差异明细 ===")
    for e in errors:
        print(f"Row {e['row']}: 工单={e['order']}, 物料={e['material']}, "
              f"优先级={e['priority']}, 替代组={e['group']}, 需求={e['demand']}, "
              f"列={e['column']}: Excel={e['excel_value']}, 计算={e['computed_value']}, "
              f"差异={e['diff']}")
else:
    print("所有分配结果完全正确！")
