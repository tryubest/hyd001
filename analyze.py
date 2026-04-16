import openpyxl
from collections import defaultdict

EXCEL_PATH = r'c:\Users\huangyuandong\Desktop\EXPORT_20260416_161509.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

rows_data = []
for row_idx in range(2, ws.max_row + 1):
    a_val = ws.cell(row_idx, 1).value
    b_val = ws.cell(row_idx, 2).value
    order_id = str(a_val).strip() if a_val and str(a_val).strip() else str(b_val).strip()
    row = {
        'row_idx': row_idx,
        'order': order_id,
        'material': ws.cell(row_idx, 3).value,
        'priority': int(ws.cell(row_idx, 4).value),
        'demand': float(ws.cell(row_idx, 6).value or 0),
        'group': str(ws.cell(row_idx, 7).value),
        'prob': float(ws.cell(row_idx, 8).value or 0),
        'init_inv': float(ws.cell(row_idx, 9).value or 0),
        'init_po': float(ws.cell(row_idx, 10).value or 0),
        'excel_L': float(ws.cell(row_idx, 12).value or 0),
        'excel_M': float(ws.cell(row_idx, 13).value or 0),
        'excel_N': ws.cell(row_idx, 14).value or '',
        'excel_O': float(ws.cell(row_idx, 15).value or 0),
        'excel_P': float(ws.cell(row_idx, 16).value or 0),
        'excel_Q': float(ws.cell(row_idx, 17).value or 0),
    }
    rows_data.append(row)

mat_initial_inv = {}
mat_initial_po = {}
for r in rows_data:
    mat = r['material']
    if r['init_inv'] > 0 and mat not in mat_initial_inv:
        mat_initial_inv[mat] = r['init_inv']
    if r['init_po'] > 0 and mat not in mat_initial_po:
        mat_initial_po[mat] = r['init_po']

mat_remaining_inv = dict(mat_initial_inv)
mat_remaining_po = dict(mat_initial_po)

order_groups = []
i = 0
while i < len(rows_data):
    current_order = rows_data[i]['order']
    current_group = rows_data[i]['group']
    group_rows = []
    while i < len(rows_data) and rows_data[i]['order'] == current_order and rows_data[i]['group'] == current_group:
        group_rows.append(rows_data[i])
        i += 1
    order_groups.append(group_rows)

TRACE_CASES = [
    ('828549', '27'),   # 案例1: 库存不够用PO补
    ('828549', '23'),   # 案例2: 优先级1有库存，不够再从优先级2取
    ('828551', '12'),   # 案例3: 有欠数的情况
    ('1200200893', 'A1'),  # 案例4: 最简单的情况
]

trace_set = set(TRACE_CASES)
errors = []
total_groups = 0
correct_groups = 0
case_num = 0

for grp_rows in order_groups:
    total_groups += 1
    grp_rows_sorted = sorted(grp_rows, key=lambda x: x['priority'])
    order_id = grp_rows_sorted[0]['order']
    group_id = grp_rows_sorted[0]['group']

    is_trace = (order_id, group_id) in trace_set
    if is_trace:
        case_num += 1
        print(f"{'='*70}")
        print(f"【追踪案例 {case_num}】工单={order_id}, 替代料组={group_id}")
        print(f"{'='*70}")

    total_demand = sum(r['demand'] for r in grp_rows_sorted)

    if is_trace:
        print(f"\n--- 步骤1: 识别该组物料 ---")
        for r in grp_rows_sorted:
            mat = r['material']
            inv_before = mat_remaining_inv.get(mat, 0)
            po_before = mat_remaining_po.get(mat, 0)
            print(f"  Row{r['row_idx']}: 物料={mat}, 优先级={r['priority']}, "
                  f"需求={r['demand']:,.0f}, "
                  f"初始库存I={r['init_inv']:,.0f}, 初始PO_J={r['init_po']:,.0f}, "
                  f"当前剩余库存={inv_before:,.0f}, 当前剩余PO={po_before:,.0f}")
        print(f"  >>> 总需求 = {total_demand:,.0f}")

    remaining_demand = total_demand
    computed_L = {}
    computed_M = {}

    if is_trace:
        print(f"\n--- 步骤2: 按优先级从低到高分配库存 ---")

    for r in grp_rows_sorted:
        mat = r['material']
        avail = mat_remaining_inv.get(mat, 0)
        alloc = min(avail, remaining_demand)
        computed_L[r['row_idx']] = alloc
        if is_trace:
            print(f"  优先级{r['priority']} 物料{mat}: "
                  f"可用库存={avail:,.0f}, 分配={alloc:,.0f}, "
                  f"剩余需求={remaining_demand:,.0f} -> {remaining_demand - alloc:,.0f}")
        remaining_demand -= alloc

    if is_trace:
        print(f"  >>> 库存分配后剩余需求 = {remaining_demand:,.0f}")
        print(f"\n--- 步骤3: 按优先级从低到高分配PO ---")

    for r in grp_rows_sorted:
        mat = r['material']
        avail = mat_remaining_po.get(mat, 0)
        alloc = min(avail, remaining_demand)
        computed_M[r['row_idx']] = alloc
        if is_trace:
            print(f"  优先级{r['priority']} 物料{mat}: "
                  f"可用PO={avail:,.0f}, 分配={alloc:,.0f}, "
                  f"剩余需求={remaining_demand:,.0f} -> {remaining_demand - alloc:,.0f}")
        remaining_demand -= alloc

    shortage = remaining_demand
    if is_trace:
        print(f"  >>> PO分配后剩余需求(欠数) = {shortage:,.0f}")

    for r in grp_rows_sorted:
        mat = r['material']
        mat_remaining_inv[mat] = mat_remaining_inv.get(mat, 0) - computed_L[r['row_idx']]
        mat_remaining_po[mat] = mat_remaining_po.get(mat, 0) - computed_M[r['row_idx']]

    if is_trace:
        print(f"\n--- 步骤4: 计算结果 vs Excel对比 ---")
        print(f"  {'列':<12} {'物料':<20} {'优先级':<6} {'计算值':>12} {'Excel值':>12} {'是否一致':>8}")
        print(f"  {'-'*72}")

    has_error = False
    for r in grp_rows_sorted:
        rid = r['row_idx']
        mat = r['material']
        computed_O = mat_remaining_inv.get(mat, 0)
        computed_P = mat_remaining_po.get(mat, 0)
        is_min_pri = r['priority'] == min(x['priority'] for x in grp_rows_sorted)
        computed_Q = shortage if is_min_pri else 0

        checks = [
            ('L(占用库存)', computed_L[rid], r['excel_L']),
            ('M(占用PO)', computed_M[rid], r['excel_M']),
            ('O(剩余库存)', computed_O, r['excel_O']),
            ('P(剩余PO)', computed_P, r['excel_P']),
            ('Q(净需求)', computed_Q, r['excel_Q']),
        ]
        for col_name, comp_val, excel_val in checks:
            match = abs(comp_val - excel_val) < 0.01
            if is_trace:
                flag = "OK" if match else "!!! 差异 !!!"
                print(f"  {col_name:<12} {mat:<20} {r['priority']:<6} {comp_val:>12,.0f} {excel_val:>12,.0f} {flag:>8}")
            if not match:
                has_error = True
                errors.append({
                    'row': rid, 'order': order_id, 'material': mat,
                    'priority': r['priority'], 'group': group_id,
                    'column': col_name, 'excel': excel_val, 'computed': comp_val,
                })

    if not has_error:
        correct_groups += 1

    if is_trace:
        print()

print(f"\n{'='*70}")
print(f"【最终验证汇总】")
print(f"{'='*70}")
print(f"总验证组数: {total_groups}")
print(f"正确组数:   {correct_groups}")
print(f"差异组数:   {total_groups - correct_groups}")
print(f"差异明细:   {len(errors)} 条")

if errors:
    print(f"\n差异详情:")
    for e in errors:
        print(f"  Row{e['row']}: 工单={e['order']} 物料={e['material']} "
              f"优先级={e['priority']} 组={e['group']} "
              f"{e['column']}: 计算={e['computed']:,.0f} Excel={e['excel']:,.0f}")
else:
    print("\n>>> 全部170个工单-替代料组验证通过，无任何差异 <<<")
